"""Write TrailO answer protocol to a native .xlsx workbook."""

from __future__ import annotations

import logging
import os
import re
from html import unescape
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.worksheet.worksheet import Worksheet

from sportorg.modules.reports.trailo_protocol import (
    ProtocolBlock,
    ProtocolField,
    SplitCell,
    TrailoMode,
    TrailoProtocolOptions,
    build_protocol_blocks,
    group_header_detail_lines,
    prepare_race_dict,
)

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FONT = Font(bold=True, size=8)
DATA_FONT = Font(size=10)
DATA_FONT_NARROW = Font(size=9)
_NARROW_PRINT_COLUMN_KEYS = frozenset({"year", "qual"})
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)
FILL_CORRECT = PatternFill("solid", fgColor="C6EFCE")
FILL_INCORRECT = PatternFill("solid", fgColor="FFC7CE")
FILL_INCORRECT_CELL = PatternFill("solid", fgColor="FFFDE8")

# Approximate print layout from 1_results_trailo.html (px → Excel width).
_COLUMN_WIDTH_SCALE = 1.12
_COMPACT_COLUMN_SCALE = 1.0
_COMPACT_COLUMN_KEYS = frozenset(
    {
        "index",
        "year",
        "qual",
        "bib",
        "preo_pass_time",
        "preo_team_pass_time",
        "trailo_score",
        "trailo_time",
        "result",
        "place_show",
    }
)
_EXCEL_OMIT_FIELD_KEYS = frozenset({"trailo_score_penalty"})

_BASE_COLUMN_WIDTHS: Dict[str, float] = {
    "index": 3.5,
    "group": 12.0,
    "name": 34.0,
    "org": 16.0,
    "year": 9.0,
    "qual": 5.2,
    "bib": 5.5,
    "preo_pass_time": 7.5,
    "preo_team_pass_time": 7.2,
    "trailo_score": 7.5,
    "trailo_time": 7.5,
    "result": 7.5,
    "result_relay": 8.5,
    "place_show": 5.2,
}
_DEFAULT_COLUMN_WIDTH = 6.8
_ANSWER_COLUMN_WIDTH = 2.5
_TIME_COLUMN_WIDTH = 6.0
_ROW_HEIGHT_DATA = 13.5
_ROW_HEIGHT_TABLE_HEADER = 24.0
_ROW_HEIGHT_GROUP_TITLE = 10.0
_ROW_HEIGHT_GROUP_META = 12.0
_ROW_HEIGHT_SIGNATURE = 14.0
_SIGNATURE_GAP_ROWS = 3
_SIGNATURE_GAP_ROW_HEIGHT = 16.0
_SIGNATURE_FONT_SIZE = 9
_SIGNATURE_IMAGE_MAX_WIDTH = 140
_SIGNATURE_IMAGE_MAX_HEIGHT = 52
_SIGNATURE_ROW_HEIGHT_WITH_IMAGE = 42.0
# Non-breaking spaces reserve room for a signature scan (no underline characters).
_SIGNATURE_FIELD_GAP = "\u00a0" * 80
_SIGNATURE_LABEL_X_OFFSET_PX = 72
_SIGNATURE_IMAGE_Y_OFFSET_PX = 6
_GROUP_META_FONT_SIZE = 10
_RACE_DESCRIPTION_FONT_SIZE = 9
_RACE_TITLE_FONT_SIZE = 14
_RACE_META_FONT_SIZE = 9
_TABLE_HEADER_FONT_SIZE = 8
_TABLE_HEADER_MAX_LINES = 2


def _set_row_height(ws: Worksheet, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def _row_height_for_wrapped_lines(line_count: int, font_size: int) -> float:
    return max(12.0, line_count * (font_size * 1.45) + 6.0)


def _data_font_for_field(field: Optional[ProtocolField]) -> Font:
    if field and field.key in _NARROW_PRINT_COLUMN_KEYS:
        return DATA_FONT_NARROW
    return DATA_FONT


def _header_title_for_field(field: ProtocolField) -> str:
    if field.key == "name":
        return "Фамилия Имя Отчество"
    if field.key == "year":
        return "Дата\nрождения"
    if field.key == "preo_pass_time":
        return "Время"
    if field.key == "preo_team_pass_time":
        return "Время"
    if field.key == "trailo_score":
        return "Результат\nочки"
    if field.key in ("trailo_time", "result"):
        return "Результат\nВремя"
    if field.key == "result_relay":
        return "Результат\nкоманды"
    title = field.title
    if len(title) > 10 and "\n" not in title:
        words = title.split()
        if len(words) >= 2:
            mid = (len(words) + 1) // 2
            return "\n".join([" ".join(words[:mid]), " ".join(words[mid:])])
    return title


def _official_names(race: Dict[str, Any]) -> Tuple[str, str]:
    data = race.get("data") or {}
    chief = str(data.get("chief_referee") or "").strip()
    secretary = str(data.get("secretary") or "").strip()
    return chief, secretary


def _signature_image_paths(race: Dict[str, Any]) -> Tuple[str, str]:
    data = race.get("data") or {}
    chief_path = str(data.get("chief_referee_signature_path") or "").strip()
    secretary_path = str(data.get("secretary_signature_path") or "").strip()
    if chief_path and not os.path.isfile(chief_path):
        chief_path = ""
    if secretary_path and not os.path.isfile(secretary_path):
        secretary_path = ""
    return chief_path, secretary_path


def _place_signature_image(
    ws: Worksheet,
    row: int,
    anchor_col: int,
    image_path: str,
    *,
    x_offset_px: int = 0,
    y_offset_px: int = _SIGNATURE_IMAGE_Y_OFFSET_PX,
    max_width: int = _SIGNATURE_IMAGE_MAX_WIDTH,
    max_height: int = _SIGNATURE_IMAGE_MAX_HEIGHT,
) -> bool:
    if not image_path:
        return False
    try:
        image = XLImage(image_path)
    except ImportError:
        logger.warning("Pillow is required to embed signature images in Excel")
        return False
    width_ratio = max_width / float(image.width or 1)
    height_ratio = max_height / float(image.height or 1)
    scale = min(width_ratio, height_ratio, 1.0)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    marker = AnchorMarker(
        col=max(anchor_col - 1, 0),
        row=max(row - 1, 0),
        colOff=pixels_to_EMU(x_offset_px),
        rowOff=pixels_to_EMU(y_offset_px),
    )
    image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(
            pixels_to_EMU(image.width),
            pixels_to_EMU(image.height),
        ),
    )
    ws.add_image(image)
    return True


def _unique_worksheet_title(wb: Workbook, base_name: str) -> str:
    title = re.sub(r"[:\\/?*\[\]]+", "_", str(base_name or "").strip()) or "Group"
    title = title[:31]
    if title not in wb.sheetnames:
        return title
    for index in range(2, 1000):
        suffix = f"_{index}"
        candidate = f"{title[: 31 - len(suffix)]}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
    return "Group"


def _prepare_worksheet(wb: Workbook, title: str, *, is_first: bool) -> Worksheet:
    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)
    ws.sheet_format.defaultRowHeight = _ROW_HEIGHT_DATA
    ws.sheet_format.customHeight = True
    return ws


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, SplitCell):
        return value.text
    return str(value)


def _strip_html(text: str) -> str:
    plain = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    plain = re.sub(r"<[^>]+>", "", plain)
    return unescape(plain).strip()


def _is_description_continuation(previous: str, line: str) -> bool:
    """True when ``line`` continues the previous wrapped description line."""
    chunk = line.strip()
    if not chunk:
        return False
    if chunk[0].islower():
        return True
    prev = previous.rstrip()
    if not prev:
        return False
    if prev.endswith(("-", "–", "—")):
        return True
    for suffix in (
        " с",
        " и",
        " к",
        " о",
        " в",
        " на",
        " по",
        " за",
        " от",
        " до",
        " при",
        " для",
        " об",
        " из",
    ):
        if prev.endswith(suffix):
            return True
    if prev.count("«") > prev.count("»"):
        return True
    if prev.count("(") > prev.count(")"):
        return True
    return False


def _normalize_description_lines(description: str) -> List[str]:
    """Merge soft line wraps so each organization is one Excel row."""
    parts = [part.strip() for part in description.splitlines() if part.strip()]
    if not parts:
        return []
    merged: List[str] = [parts[0]]
    for part in parts[1:]:
        if _is_description_continuation(merged[-1], part):
            merged[-1] = f"{merged[-1]} {part}"
        else:
            merged.append(part)
    return merged


def _race_date_place_line(race: Dict[str, Any]) -> str:
    data = race.get("data") or {}
    when = str(data.get("start_datetime") or "")[:10]
    if when and re.match(r"^\d{4}-\d{2}-\d{2}$", when):
        parts = when.split("-")
        when = f"{parts[2]}.{parts[1]}.{parts[0]}"
    where = str(data.get("location") or "").strip()
    return ", ".join(part for part in (when, where) if part)


def _write_merged_header_line(
    ws: Worksheet,
    row: int,
    col_count: int,
    text: str,
    *,
    font: Font,
) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    size = int(font.size or _RACE_DESCRIPTION_FONT_SIZE)
    _set_row_height(ws, row, _row_height_for_wrapped_lines(1, size))
    return row + 1


def _write_header_block(
    ws: Worksheet,
    row: int,
    col_count: int,
    race: Dict[str, Any],
    *,
    show_answers: bool,
) -> int:
    """Description: one row per organization, 9 pt. Title: 14 pt bold."""
    data = race.get("data") or {}
    description = _strip_html(str(data.get("description") or ""))
    if description:
        desc_font = Font(size=_RACE_DESCRIPTION_FONT_SIZE)
        for line in _normalize_description_lines(description):
            row = _write_merged_header_line(
                ws, row, col_count, line, font=desc_font
            )
    title = str(data.get("title") or "").strip()
    if title:
        row = _write_merged_header_line(
            ws,
            row,
            col_count,
            title,
            font=Font(bold=True, size=_RACE_TITLE_FONT_SIZE),
        )
    date_place = _race_date_place_line(race)
    meta_font = Font(bold=True, size=_RACE_META_FONT_SIZE)
    if date_place:
        row = _write_merged_header_line(ws, row, col_count, date_place, font=meta_font)
    protocol_label = (
        "Предварительный протокол ответов"
        if show_answers
        else "Протокол результатов"
    )
    row = _write_merged_header_line(
        ws, row, col_count, protocol_label, font=meta_font
    )
    return row


def _signature_row_text(label: str, name: str) -> str:
    if name:
        return f"{label}{_SIGNATURE_FIELD_GAP}{name}"
    return f"{label}{_SIGNATURE_FIELD_GAP}"


def _signature_image_x_offset(label: str) -> int:
    """Place the scan to the right of the duty title (~9pt Calibri in Excel)."""
    return int(len(label) * 10) + _SIGNATURE_LABEL_X_OFFSET_PX


def _write_official_signature_row(
    ws: Worksheet,
    row: int,
    col_count: int,
    label: str,
    name: str,
    image_path: str,
    *,
    signature_x_offset_px: int,
) -> int:
    """One row: label, room for signature (optional image), official name."""
    sig_font = Font(size=_SIGNATURE_FONT_SIZE)
    line = _signature_row_text(label, name)
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=col_count,
    )
    cell = ws.cell(row=row, column=1, value=line)
    cell.font = sig_font
    cell.alignment = LEFT_NOWRAP
    has_image = False
    if image_path:
        has_image = _place_signature_image(
            ws,
            row,
            1,
            image_path,
            x_offset_px=signature_x_offset_px,
        )
    row_height = (
        _SIGNATURE_ROW_HEIGHT_WITH_IMAGE if has_image else _ROW_HEIGHT_SIGNATURE
    )
    _set_row_height(ws, row, row_height)
    return row + 1


def _write_signature_block(
    ws: Worksheet, row: int, col_count: int, race: Dict[str, Any]
) -> int:
    """Chief referee and secretary on separate rows after the results table."""
    chief, secretary = _official_names(race)
    chief_image, secretary_image = _signature_image_paths(race)
    for _ in range(_SIGNATURE_GAP_ROWS):
        _set_row_height(ws, row, _SIGNATURE_GAP_ROW_HEIGHT)
        row += 1
    chief_label = "Главный судья"
    secretary_label = "Главный секретарь"
    row = _write_official_signature_row(
        ws,
        row,
        col_count,
        chief_label,
        chief,
        chief_image,
        signature_x_offset_px=_signature_image_x_offset(chief_label),
    )
    row = _write_official_signature_row(
        ws,
        row,
        col_count,
        secretary_label,
        secretary,
        secretary_image,
        signature_x_offset_px=_signature_image_x_offset(secretary_label),
    )
    return row


def _fields_for_excel(fields: List[ProtocolField]) -> List[ProtocolField]:
    return [field for field in fields if field.key not in _EXCEL_OMIT_FIELD_KEYS]


def _active_field_keys(fields: List[ProtocolField]) -> List[str]:
    return [field.key for field in fields if field.active]


def _field_by_key(fields: List[ProtocolField]) -> Dict[str, ProtocolField]:
    return {field.key: field for field in fields}


def _split_station_groups(
    fields: List[ProtocolField], mode: TrailoMode
) -> Tuple[List[ProtocolField], List[List[ProtocolField]], List[List[ProtocolField]]]:
    """Return base fields, main-course answer fields, and station field groups."""
    base: List[ProtocolField] = []
    answers: List[ProtocolField] = []
    stations: List[List[ProtocolField]] = []
    current_station: List[ProtocolField] = []
    time_title = mode.time_control_title()
    for field in fields:
        if re.match(r"^\d+_", field.key):
            if field.is_time or field.title == time_title:
                if current_station:
                    stations.append(current_station)
                current_station = [field]
            else:
                if current_station:
                    current_station.append(field)
                else:
                    answers.append(field)
        else:
            base.append(field)
    if current_station:
        stations.append(current_station)
    return base, answers, stations


def _write_group_block(
    ws: Worksheet,
    block: ProtocolBlock,
    race: Dict[str, Any],
    mode: TrailoMode,
    *,
    show_answers: bool,
) -> Tuple[int, int]:
    """Write race header, group title, table headers and data. Returns (last_row, col_count)."""
    fields = _fields_for_excel(block.fields)
    col_count = max(len(_active_field_keys(fields)), 1)
    _apply_column_widths(ws, fields)
    row = 1
    row = _write_header_block(ws, row, col_count, race, show_answers=show_answers)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    title_cell = ws.cell(row=row, column=1, value=block.name)
    title_cell.font = Font(bold=True, size=10)
    title_cell.alignment = CENTER
    _set_row_height(ws, row, _ROW_HEIGHT_GROUP_TITLE)
    row += 1
    for line in group_header_detail_lines(block.group, block.course, mode):
        row = _write_merged_header_line(
            ws,
            row,
            col_count,
            line,
            font=Font(size=_GROUP_META_FONT_SIZE),
        )
        _set_row_height(ws, row - 1, _ROW_HEIGHT_GROUP_META)
    row = _write_table_headers(ws, row, fields, mode, block.hide_answer_labels)
    row = _write_data_rows(ws, row, fields, block.rows)
    row = _write_signature_block(ws, row, col_count, race)
    return row, col_count


def _write_table_headers(
    ws: Worksheet,
    row: int,
    fields: List[ProtocolField],
    mode: TrailoMode,
    hide_answer_labels: bool,
) -> int:
    base_fields, answer_fields, station_groups = _split_station_groups(fields, mode)
    has_split_columns = bool(answer_fields or station_groups)
    if not has_split_columns:
        header_rows = 1
        sub_header_rowspan = 1
    else:
        header_rows = 2 if hide_answer_labels else 3
        sub_header_rowspan = 1 if hide_answer_labels else 2
    col = 1

    def write_cell(
        r: int,
        c: int,
        value: str,
        *,
        rowspan: int = 1,
        colspan: int = 1,
        font: Optional[Font] = None,
    ) -> None:
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font or HEADER_FONT
        cell.alignment = HEADER_WRAP
        cell.border = THIN_BORDER
        if rowspan > 1 or colspan > 1:
            ws.merge_cells(
                start_row=r,
                start_column=c,
                end_row=r + rowspan - 1,
                end_column=c + colspan - 1,
            )

    for field in base_fields:
        title = _header_title_for_field(field)
        write_cell(row, col, title, rowspan=header_rows)
        col += 1

    if answer_fields:
        write_cell(row, col, "Ответы", colspan=len(answer_fields))
        for index, field in enumerate(answer_fields, start=1):
            write_cell(row + 1, col + index - 1, str(index))
            if not hide_answer_labels:
                write_cell(row + 2, col + index - 1, field.title)
        col += len(answer_fields)

    for station_index, station_fields in enumerate(station_groups, start=1):
        station_title = (
            f"Станция {station_index}" if mode.is_tempo else f"Тайм КП {station_index}"
        )
        write_cell(row, col, station_title, colspan=len(station_fields))
        write_cell(row + 1, col, "Время", rowspan=sub_header_rowspan)
        answer_cols = station_fields[1:]
        for index, field in enumerate(answer_cols, start=1):
            write_cell(row + 1, col + index, str(index))
            if not hide_answer_labels:
                write_cell(row + 2, col + index, field.title)
        col += len(station_fields)

    header_line_height = _row_height_for_wrapped_lines(
        _TABLE_HEADER_MAX_LINES, _TABLE_HEADER_FONT_SIZE
    )
    for header_row in range(row, row + header_rows):
        _set_row_height(ws, header_row, max(_ROW_HEIGHT_TABLE_HEADER, header_line_height))

    return row + header_rows


def _style_data_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    field: Optional[ProtocolField] = None,
    align_left: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=_cell_text(value))
    cell.border = THIN_BORDER
    if align_left:
        cell.alignment = LEFT_NOWRAP if field and field.key == "name" else LEFT
    else:
        cell.alignment = CENTER
    cell.font = _data_font_for_field(field)
    if isinstance(value, SplitCell):
        if value.is_answer and value.is_correct is True:
            cell.font = Font(size=10, color="006100", bold=True)
            cell.fill = FILL_CORRECT
        elif value.is_answer and value.is_correct is False:
            cell.font = Font(size=10, color="9C0006", bold=True)
            cell.fill = FILL_INCORRECT_CELL


def _write_data_rows(
    ws: Worksheet,
    start_row: int,
    fields: List[ProtocolField],
    rows: List[Dict[str, Any]],
) -> int:
    keys = _active_field_keys(fields)
    field_map = _field_by_key(fields)
    row_idx = start_row
    merge_ranges: List[Tuple[int, int, int, int]] = []
    for data_row in rows:
        col = 1
        rowspan = int(data_row.get("_relay_rowspan") or 1)
        skip_fields = set(data_row.get("_relay_skip_fields") or [])
        merged_fields = set(data_row.get("_relay_merged_fields") or [])
        for key in keys:
            field = field_map.get(key)
            value = data_row.get(key, "")
            if key not in skip_fields:
                align_left = key in ("name", "org", "group")
                _style_data_cell(
                    ws, row_idx, col, value, field=field, align_left=align_left
                )
                if rowspan > 1 and key in merged_fields:
                    merge_ranges.append((row_idx, col, row_idx + rowspan - 1, col))
            col += 1
        _set_row_height(ws, row_idx, _ROW_HEIGHT_DATA)
        row_idx += 1
    for start_r, start_c, end_r, end_c in merge_ranges:
        if end_r > start_r:
            ws.merge_cells(
                start_row=start_r,
                start_column=start_c,
                end_row=end_r,
                end_column=end_c,
            )
            ws.cell(row=start_r, column=start_c).alignment = CENTER
    return row_idx


def _column_width_for_field(field: ProtocolField) -> float:
    if re.match(r"^\d+_", field.key):
        width = _TIME_COLUMN_WIDTH if field.is_time else _ANSWER_COLUMN_WIDTH
        scale = _COLUMN_WIDTH_SCALE
    else:
        width = _BASE_COLUMN_WIDTHS.get(field.key, _DEFAULT_COLUMN_WIDTH)
        scale = (
            _COMPACT_COLUMN_SCALE
            if field.key in _COMPACT_COLUMN_KEYS
            else _COLUMN_WIDTH_SCALE
        )
    return round(width * scale, 1)


def _apply_column_widths(ws: Worksheet, fields: List[ProtocolField]) -> None:
    active_fields = [field for field in fields if field.active]
    for col_idx, field in enumerate(active_fields, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _column_width_for_field(field)


def default_excel_protocol_options() -> TrailoProtocolOptions:
    """Excel export: results table only (no answer / station columns)."""
    return TrailoProtocolOptions(show_answers=False)


def _apply_print_setup(
    ws: Worksheet,
    last_row: int,
    last_col: int,
    race: Dict[str, Any],
    *,
    show_answers: bool,
) -> None:
    """A4: portrait for results-only, landscape when answer columns are included."""
    if show_answers:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    else:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1
    if last_row > 0 and last_col > 0:
        ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"


def save_trailo_protocol_excel(
    race: Dict[str, Any],
    file_name: str,
    *,
    options: Optional[TrailoProtocolOptions] = None,
) -> None:
    if options is None:
        options = default_excel_protocol_options()
    race = prepare_race_dict(race)
    mode = TrailoMode(race)
    blocks = build_protocol_blocks(race, options)
    wb = Workbook()
    show_answers = options.show_answers
    if not blocks:
        ws = wb.active
        ws.title = "Протокол"
        row = 1
        row = _write_header_block(ws, row, 1, race, show_answers=show_answers)
        ws.cell(row=row, column=1, value="Нет данных для протокола")
        wb.save(file_name)
        return

    sheet_per_group = len(blocks) > 1

    for block_index, block in enumerate(blocks):
        if sheet_per_group:
            sheet_title = _unique_worksheet_title(
                wb, block.short_name or block.name
            )
            ws = _prepare_worksheet(wb, sheet_title, is_first=block_index == 0)
        else:
            ws = wb.active
            ws.title = "Протокол"
        row, col_count = _write_group_block(
            ws, block, race, mode, show_answers=show_answers
        )
        _apply_print_setup(
            ws,
            last_row=max(row - 1, 1),
            last_col=col_count,
            race=race,
            show_answers=show_answers,
        )
    wb.save(file_name)


def default_excel_filename(race: Dict[str, Any]) -> str:
    race = prepare_race_dict(race)
    mode = TrailoMode(race)
    data = race.get("data") or {}
    desc = str(data.get("description") or "results")
    base = re.sub(r'[\\/:*?"<>|]+', "_", desc)
    base = re.sub(r"\s+", "_", base).strip("_")[:80] or "results"
    when = str(data.get("start_datetime") or "")[:10].replace("-", "")
    mode_tag = "preo" if mode.is_preo else ("tempo" if mode.is_tempo else "sprint")
    suffix = f"_{when}" if when else ""
    return f"trailo_protocol_{mode_tag}_{base}{suffix}"
