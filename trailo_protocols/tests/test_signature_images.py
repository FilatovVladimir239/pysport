import base64
import os
import tempfile

import pytest
from openpyxl import load_workbook

try:
    from PIL import Image as PILImage  # noqa: F401

    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False

from trailo_protocols.excel import save_trailo_protocol_excel
from trailo_protocols.export_core import default_plugin_settings
from trailo_protocols.signature_images import (
    migrate_signature_settings_from_race,
    store_signature_image,
)
from fixtures_preo import setup_preo_group

_MIN_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


def test_migrate_signature_settings_from_race():
    settings = default_plugin_settings()
    race = {
        "data": {
            "chief_referee_signature_path": r"C:\legacy\chief.png",
            "secretary_signature_path": r"C:\legacy\sec.png",
        }
    }
    merged = migrate_signature_settings_from_race(settings, race)
    assert merged["chief_referee_signature_path"] == r"C:\legacy\chief.png"
    assert merged["secretary_signature_path"] == r"C:\legacy\sec.png"


def test_store_signature_image_copies_file():
    plugin_settings = default_plugin_settings()
    plugin_settings["output_dir"] = tempfile.mkdtemp()
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, "sign.png")
        with open(source, "wb") as handle:
            handle.write(_MIN_PNG)
        saved = store_signature_image(source, "chief_referee", plugin_settings)
        assert os.path.isfile(saved)
        assert saved != source
        assert "signatures" in saved.replace("\\", "/")


@pytest.mark.skipif(not HAS_PILLOW, reason="Pillow is required for Excel signature images")
def test_excel_protocol_embeds_signature_images():
    race_dict = setup_preo_group()
    plugin_settings = default_plugin_settings()
    with tempfile.TemporaryDirectory() as tmp:
        plugin_settings["output_dir"] = tmp
        chief_path = os.path.join(tmp, "chief.png")
        secretary_path = os.path.join(tmp, "secretary.png")
        for path in (chief_path, secretary_path):
            with open(path, "wb") as handle:
                handle.write(_MIN_PNG)
        plugin_settings["chief_referee_signature_path"] = chief_path
        plugin_settings["secretary_signature_path"] = secretary_path
        out = os.path.join(tmp, "protocol.xlsx")
        save_trailo_protocol_excel(race_dict, out, plugin_settings=plugin_settings)
        wb = load_workbook(out)
        assert len(wb.active._images) >= 2


@pytest.mark.skipif(not HAS_PILLOW, reason="Pillow is required for Excel signature images")
def test_excel_protocol_embeds_federation_stamp():
    race_dict = setup_preo_group()
    plugin_settings = default_plugin_settings()
    with tempfile.TemporaryDirectory() as tmp:
        plugin_settings["output_dir"] = tmp
        stamp_path = os.path.join(tmp, "federation.png")
        with open(stamp_path, "wb") as handle:
            handle.write(_MIN_PNG)
        plugin_settings["federation_stamp_path"] = stamp_path
        out = os.path.join(tmp, "protocol_stamp.xlsx")
        save_trailo_protocol_excel(race_dict, out, plugin_settings=plugin_settings)
        wb = load_workbook(out)
        assert len(wb.active._images) >= 1
