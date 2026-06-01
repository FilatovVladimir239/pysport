import os
import tempfile

import pytest
from openpyxl import load_workbook

from trailo_protocols.excel_script import (
    resolve_excel_script_path,
    run_excel_export_script,
)
from fixtures_preo import setup_preo_group


def test_resolve_excel_script_path_relative():
    root = os.path.join(tempfile.gettempdir(), "sportorg_templates")
    path = resolve_excel_script_path("/reports/custom.py", root)
    assert path == os.path.normpath(os.path.join(root, "reports", "custom.py"))


def test_run_excel_export_script_calls_export():
    race = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        script = os.path.join(tmp, "export_protocol.py")
        out = os.path.join(tmp, "out.xlsx")
        with open(script, "w", encoding="utf-8") as handle:
            handle.write(
                "def export(race, file_name, show_answers=False):\n"
                "    from openpyxl import Workbook\n"
                "    wb = Workbook()\n"
                "    wb.active['A1'] = race['persons'][0]['surname']\n"
                "    wb.save(file_name)\n"
            )
        run_excel_export_script(script, race, out, show_answers=False)
        wb = load_workbook(out)
        assert wb.active["A1"].value == "Ivanov"


def test_run_excel_export_script_missing_entry_raises():
    race = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        script = os.path.join(tmp, "empty.py")
        with open(script, "w", encoding="utf-8") as handle:
            handle.write("value = 1\n")
        with pytest.raises(AttributeError):
            run_excel_export_script(script, race, os.path.join(tmp, "out.xlsx"))
