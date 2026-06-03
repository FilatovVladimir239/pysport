import os
import re
import tempfile
from openpyxl import load_workbook

from sportorg.common.otime import OTime
from sportorg.models.memory import (
    Course,
    CourseControl,
    Group,
    Organization,
    Person,
    Qualification,
    Race,
    RaceType,
    ResultManual,
    ResultStatus,
    Split,
    new_event,
    race,
)
from sportorg.modules.trailo.result_checker import TrailoResultChecker
from sportorg.modules.reports.trailo_protocol import (
    SplitCell,
    TrailoMode,
    TrailoProtocolOptions,
    build_base_fields,
    build_protocol_blocks,
    count_trailo_course_controls,
    group_display_name,
    group_header_detail_lines,
    prepare_race_dict,
)
from trailo_protocols.excel import (
    _normalize_description_lines,
    default_excel_protocol_options,
    save_trailo_protocol_excel,
)
from sportorg.models.result.result_tools import recalculate_results
from fixtures_preo import setup_preo_group


def test_build_protocol_blocks_preo():
    race_dict = prepare_race_dict(setup_preo_group())
    blocks = build_protocol_blocks(race_dict)
    assert len(blocks) == 1
    block = blocks[0]
    assert block.name == "Мужчины 21"
    assert block.short_name == "M21"
    assert len(block.rows) == 2
    ok_row = next(row for row in block.rows if row.get("bib") == 101)
    assert ok_row["name"] == "Ivanov Ivan"
    assert isinstance(ok_row.get("0_31"), SplitCell)


def test_count_trailo_course_controls_preo():
    race_dict = prepare_race_dict(setup_preo_group())
    mode = TrailoMode(race_dict)
    course = race_dict["groups"][0]["course"]
    main_count, time_tc_count = count_trailo_course_controls(course, mode)
    assert main_count == 1
    assert time_tc_count == 1


def test_group_header_detail_lines_preo():
    race_dict = prepare_race_dict(setup_preo_group())
    mode = TrailoMode(race_dict)
    group = race_dict["groups"][0]
    course = group["course"]
    lines = group_header_detail_lines(group, course, mode)
    assert any("Контрольное время — 00:45:00" in line for line in lines)
    assert any(
        "Дистанция: КМ —" in line and ", КП — 1," in line and "тайм КП — 1" in line
        for line in lines
    )


def _setup_tempo_group():
    new_event([Race()])
    race().data.race_type = RaceType.INDIVIDUAL_RACE
    race().set_setting("result_processing_mode", "trailo")
    race().set_setting("trailo_mode", "tempo")
    race().set_setting("trailo_alternate_course", False)

    group = Group()
    group.name = "Open"
    course = Course()
    for code in ("31A", "1TT", "1T1A", "2TT", "2T1A"):
        control = CourseControl()
        control.code = code
        course.controls.append(control)
    group.course = course
    race().groups.append(group)
    race().courses.append(course)
    return race().to_dict()


def test_group_header_detail_lines_tempo():
    race_dict = prepare_race_dict(_setup_tempo_group())
    mode = TrailoMode(race_dict)
    group = race_dict["groups"][0]
    course = group["course"]
    lines = group_header_detail_lines(group, course, mode)
    assert any(line == "Дистанция: 2 станции" for line in lines)
    assert not any("КМ —" in line for line in lines)
    assert not any("тайм КП" in line for line in lines)


def _setup_tempo_group_with_results():
    _setup_tempo_group()
    group = race().groups[0]
    org = Organization()
    org.name = "Club"

    person_ok = Person()
    person_ok.group = group
    person_ok.organization = org
    person_ok.set_bib(1)
    person_ok.surname = "Fast"
    person_ok.name = "Runner"
    person_ok.qual = Qualification.III

    split = Split()
    split.code = "31A"
    split.is_correct = True
    split.time = OTime(msec=30_000)
    split.course_index = 0

    result_ok = ResultManual()
    result_ok.person = person_ok
    result_ok.status = ResultStatus.OK
    result_ok.start_time = OTime(msec=0)
    result_ok.finish_time = OTime(msec=30_000)
    result_ok.splits = [split]
    TrailoResultChecker.process(result_ok, lambda *args, **kwargs: 0)

    person_dns = Person()
    person_dns.group = group
    person_dns.organization = org
    person_dns.set_bib(2)
    person_dns.surname = "Late"
    person_dns.name = "Start"
    result_dns = ResultManual()
    result_dns.person = person_dns
    result_dns.status = ResultStatus.DID_NOT_START
    TrailoResultChecker.process(result_dns, lambda *args, **kwargs: 0)

    race().persons.extend([person_ok, person_dns])
    race().results.extend([result_ok, result_dns])
    recalculate_results(recheck_results=True)
    return race().to_dict()


def test_save_trailo_protocol_excel_tempo_result_header_and_status():
    race_dict = prepare_race_dict(_setup_tempo_group_with_results())
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "tempo.xlsx")
        save_trailo_protocol_excel(race_dict, path)
        ws = load_workbook(path).active
        flat = [str(c.value or "") for row in ws.iter_rows() for c in row]
        assert any(v == "Результат" for v in flat)
        assert any(v == "время" for v in flat)
        assert not any(v == "Время" and "Результат" not in str(v) for v in flat if v in ("Время",))
        assert any("не старт" in v for v in flat)
        result_header_height = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value == "Результат":
                    result_header_height = ws.row_dimensions[row].height
                    break
            if result_header_height is not None:
                break
        assert result_header_height is not None
        assert result_header_height >= 16.0


def test_group_display_name_prefers_long_name():
    group = {"name": "M21", "long_name": "Мужчины 21 год и старше"}
    assert group_display_name(group) == "Мужчины 21 год и старше"
    assert group_display_name({"name": "W21", "long_name": ""}) == "W21"


def test_normalize_description_lines_merges_wrapped_organizations():
    text = (
        "Министерство спорта Российской Федерации\n"
        "Общероссийская общественная организация «Всероссийская Федерация спорта лиц с\n"
        "поражением опорно-двигательного аппарата»\n"
        "СПбРОО «Спортивная Федерация\n"
        "спорта лиц с поражением опорно-двигательного аппарата и спорта лиц с\n"
        "интеллектуальными нарушениями»\n"
        "Комитет по физической культуре и спорту Санкт-Петербурга"
    )
    lines = _normalize_description_lines(text)
    assert len(lines) == 4
    assert lines[0] == "Министерство спорта Российской Федерации"
    assert lines[1].startswith("Общероссийская")
    assert "поражением опорно-двигательного аппарата»" in lines[1]
    assert lines[2].startswith("СПбРОО")
    assert "интеллектуальными нарушениями»" in lines[2]
    assert lines[3] == "Комитет по физической культуре и спорту Санкт-Петербурга"


def test_build_base_fields_result_column_titles():
    race_dict = prepare_race_dict(setup_preo_group())
    race_dict["settings"]["trailo_custom_penalty_time_enabled"] = False
    mode = TrailoMode(race_dict)
    titles = [field.title for field in build_base_fields(mode)]
    assert "Результат очки" in titles
    assert "Результат Время" in titles
    assert "Время" in titles


def test_excel_default_options_exclude_answer_columns():
    race_dict = prepare_race_dict(setup_preo_group())
    options = default_excel_protocol_options()
    assert options.show_answers is False
    blocks = build_protocol_blocks(race_dict, options)
    keys = [field.key for field in blocks[0].fields]
    assert not any(re.match(r"^\d+_", key) for key in keys)


def test_save_trailo_protocol_excel_file():
    race_dict = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "protocol.xlsx")
        save_trailo_protocol_excel(race_dict, path)
        wb = load_workbook(path)
        ws = wb.active
        values = [cell.value for row in ws.iter_rows(min_row=1, max_row=30) for cell in row]
        flat = [str(v) for v in values if v]
        assert any("Ivanov Ivan" in value for value in flat)
        assert any("Petrov Petr" in value for value in flat)
        assert any("Протокол результатов" in value for value in flat)
        assert any(value == "Результат" for value in flat)
        assert any(value == "очки" for value in flat)
        assert any(value == "время" for value in flat)
        assert not any("Предварительный протокол ответов" in value for value in flat)
        assert any("Мужчины 21" in value for value in flat)
        assert any("Контрольное время" in value for value in flat)
        assert any("Дистанция:" in value for value in flat)
        assert not any(value == "M21" for value in flat)
        assert any("Главный судья" in value and "Сидоров С.С." in value for value in flat)
        assert any(
            "Главный секретарь" in value and "Кузнецова К.К." in value for value in flat
        )
        assert not any("Станция" in value for value in flat)
        assert not any("Тайм КП" in value for value in flat)
        assert not any("Ответы" in value for value in flat)
        assert not any("Штраф" in value for value in flat)
        assert not any("Подпись" in value for value in flat)
        assert not any(value.strip() == "М.П." for value in flat)


def test_excel_merges_result_cells_for_non_ok_status():
    race_dict = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "protocol_merge.xlsx")
        save_trailo_protocol_excel(race_dict, path)
        wb = load_workbook(path)
        ws = wb.active
        # Find status cell and ensure it is merged across two columns.
        status_cell = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value == "не старт":
                    status_cell = cell
                    break
            if status_cell:
                break
        assert status_cell is not None
        merged_ranges = list(ws.merged_cells.ranges)
        assert any(status_cell.coordinate in rng for rng in merged_ranges)


def test_save_trailo_protocol_excel_a4_portrait_fit():
    race_dict = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "protocol_print.xlsx")
        save_trailo_protocol_excel(race_dict, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.page_setup.orientation == "portrait"
        assert int(ws.page_setup.paperSize) == 9
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
        values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert any("Главный судья" in value and "Сидоров С.С." in value for value in values)
        assert any(
            "Главный секретарь" in value and "Кузнецова К.К." in value for value in values
        )


def test_save_trailo_protocol_excel_page_break_and_repeated_headers():
    race_dict = prepare_race_dict(setup_preo_group())
    race_dict["settings"]["trailo_custom_penalty_time_enabled"] = False
    race_dict["data"]["title"] = "Test Event"
    race_dict["data"]["description"] = "Day 1"

    group2 = {
        "id": "g2",
        "name": "W21",
        "object": "Group",
        "course_id": race_dict["courses"][0]["id"],
        "count_finished": 0,
        "count_person": 0,
        "first_number": 0,
        "is_any_course": False,
        "is_best_team_placing_mode": False,
        "long_name": "",
        "max_age": 0,
        "max_time": 0,
        "max_year": 0,
        "min_age": 0,
        "min_year": 0,
        "order_in_corridor": 0,
        "price": 0,
        "ranking": {"is_active": False, "rank": [], "rank_scores": 0},
        "relay_legs": 0,
        "sex": 0,
        "start_corridor": 0,
        "start_interval": 0,
    }
    person2 = dict(race_dict["persons"][0])
    person2["id"] = "p2"
    person2["group_id"] = "g2"
    person2["bib"] = 102
    person2["surname"] = "Petrova"
    person2["name"] = "Anna"
    result2 = dict(race_dict["results"][0])
    result2["id"] = "r2"
    result2["person_id"] = "p2"
    race_dict["groups"].append(group2)
    race_dict["persons"].append(person2)
    race_dict["results"].append(result2)

    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "two_groups.xlsx")
        save_trailo_protocol_excel(race_dict, path)
        wb = load_workbook(path)
        ws = wb.active
        assert len(wb.sheetnames) == 2
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            flat = [
                str(cell.value)
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row)
                for cell in row
                if cell.value is not None
            ]
            assert any("Test Event" in value for value in flat)
            assert any("Day 1" in value for value in flat)
        all_flat = [
            str(cell.value)
            for sheet_name in wb.sheetnames
            for row in wb[sheet_name].iter_rows(min_row=1, max_row=wb[sheet_name].max_row)
            for cell in row
            if cell.value is not None
        ]
        assert any("Мужчины 21" in value for value in all_flat)
        assert any(value == "W21" for value in all_flat)


def test_save_trailo_protocol_excel_relay():
    import importlib.util
    from pathlib import Path

    relay_path = (
        Path(__file__).resolve().parents[2] / "tests" / "test_relay_trailo_preo.py"
    )
    spec = importlib.util.spec_from_file_location("relay_trailo_preo", relay_path)
    relay_mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(relay_mod)
    _add_leg_result = relay_mod._add_leg_result
    _setup_relay_preo = relay_mod._setup_relay_preo

    group, leg1, leg2 = _setup_relay_preo()

    split1 = Split()
    split1.code = "31A"
    split1.is_correct = True
    split1.time = OTime(msec=20_000)
    split1.course_index = 0
    split_tt1 = Split()
    split_tt1.code = "1TT"
    split_tt1.is_correct = True
    split_tt1.time = OTime(msec=40_000)
    split_tt1.course_index = 1
    split_ans1 = Split()
    split_ans1.code = "1T1A"
    split_ans1.is_correct = True
    split_ans1.time = OTime(msec=45_000)
    split_ans1.course_index = 2

    _add_leg_result(leg1, [split1, split_tt1, split_ans1], 1, 50_000)
    recalculate_results(recheck_results=True)

    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "relay.xlsx")
        save_trailo_protocol_excel(race().to_dict(), path)
        wb = load_workbook(path)
        assert wb.active.max_row >= 5
        flat = [
            str(cell.value)
            for row in wb.active.iter_rows(min_row=1, max_row=wb.active.max_row)
            for cell in row
            if cell.value is not None
        ]
        kus_line = next(
            (value for value in flat if value.startswith("Квалификационный уровень —")),
            "",
        )
        criteria = next(
            (value for value in flat if "МС — 1 место" in str(value)),
            "",
        )
        if kus_line:
            assert "МС — 1 место" in criteria
            assert "КМС — 2–3 место" in criteria
            assert "МС — 1–3 место" not in criteria


def test_save_trailo_protocol_excel_can_include_answers_when_requested():
    race_dict = setup_preo_group()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "protocol_with_answers.xlsx")
        save_trailo_protocol_excel(
            race_dict,
            path,
            options=TrailoProtocolOptions(show_answers=True),
        )
        wb = load_workbook(path)
        ws = wb.active
        flat = [
            str(cell.value)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
            for cell in row
            if cell.value is not None
        ]
        assert any("Ответы" in value or "Станция" in value or "Тайм КП" in value for value in flat)
        assert ws.page_setup.orientation == "landscape"
